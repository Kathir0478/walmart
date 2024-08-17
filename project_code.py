from openpyxl import load_workbook
import nltk
from nltk.tokenize import word_tokenize
import string
import pathlib
import textwrap

class ReviewCleaner:
    def __init__(self, stop_words_path, excel_path, positive_word_path, negative_word_path,relavant_word_path):
        nltk.download('punkt')
        self.symbols_to_remove = set(string.punctuation)
        self.stopwords_to_remove = self._load_word_data(stop_words_path, True)
        self.excel_path = excel_path
        self.positive_words = self._load_word_data(positive_word_path)
        self.negative_words = self._load_word_data(negative_word_path)
        self.relavant_words = self._load_word_data(relavant_word_path)

    def _load_word_data(self, path, is_symbol=False):
        with open(path, 'r') as file:
            words_token = set(line.strip() for line in file)
        if is_symbol:
            return words_token.union(self.symbols_to_remove)
        return words_token

    def _process_review(self, review_sentence):
        tokenized_review = word_tokenize(review_sentence)
        cleaned_words = [i.lower() for i in tokenized_review if i.lower() not in self.stopwords_to_remove]
        return cleaned_words

    def review_impression(self, review_tokens):
        impression_score = 0
        impression_balance = 0
        relavancy_score = 0
        for i in review_tokens:
            if i in self.positive_words:
                impression_score+=1
            elif i in self.negative_words:
                impression_score-=1
            if i in self.relavant_words:
                relavancy_score+=1
            else:
                impression_balance+=1
            relavancy_factor= (relavancy_score/len(review_tokens))*100
        return impression_score/impression_balance,relavancy_factor

    def output(self):
        workbook = load_workbook(self.excel_path, data_only=True)
        sheet = workbook.active
        cleaned_reviews = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            review_id = row[0]
            review_sentence = row[1]
            if review_sentence:
                cleaned_review = self._process_review(review_sentence)
                impression_score,relavancy_score = self.review_impression(cleaned_review)
                cleaned_reviews.append((review_id, cleaned_review, impression_score,relavancy_score))

        return cleaned_reviews


# Example usage:
stop_words_path = r"W:\Walmart\final\stopwords_data.txt"
excel_path = r"W:\Walmart\final\review_data.xlsx"
positive_word_path = r"W:\Walmart\final\positive_words_final.txt"
negative_word_path = r"W:\Walmart\final\negative_words_final.txt"
relavant_word_path = r"W:\Walmart\final\relavant_words_data.txt"

#review_cleaner = ReviewCleaner(stop_words_path, excel_path, positive_word_path, negative_word_path,relavant_word_path)
#cleaned_reviews = review_cleaner.output()

